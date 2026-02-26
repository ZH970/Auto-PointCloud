import argparse
import csv
import os
import re
from typing import List, Optional
import zipfile
import openpyxl
from tempfile import TemporaryDirectory
from openpyxl import load_workbook
from lxml import etree
import xml.etree.ElementTree as ET
import shutil
from pathlib import PurePosixPath
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

current_dir = os.path.dirname(os.path.abspath(__file__))
DEFAULT_CSV = os.path.join(current_dir, "points.csv")
DEFAULT_CSV1 = os.path.join(r"C:\Users\ZN191014\Desktop\新建文件夹 (2)\after\csv\1169-", "points.csv")
DEFAULT_WORKBOOK = os.path.join(current_dir, "SLAM扫描测试.xlsx")
DEFAULT_WORKBOOK1 = os.path.join(r"C:\Users\ZN191014\Documents\WXWork\1688857910833845\Cache\File\2025-11",
                                 "检测数据（新）.xlsx")
DEFAULT_WORKBOOK2 = os.path.join(r"E:\文档","检测数据.xlsx")
DEFAULT_SAVEBOOK = os.path.join(current_dir, "SLAM扫描测试.xlsx")
DEFAULT_SHEET = "SLAM扫描测试"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="将 points.csv 中的数据写入 SLAM 表格。")
    parser.add_argument("--csv", default=DEFAULT_CSV1, help="points.csv 的路径。")
    parser.add_argument("--workbook", default=DEFAULT_WORKBOOK2, help="目标 Excel 文件路径。")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="需要写入的工作表名称。")
    parser.add_argument("--preserve-image-column", type=int, default=1, help="保留图像列的列号（从1开始）。")
    parser.add_argument(
        "--save-as",
        default=os.path.join(current_dir, "SLAM扫描测试.xlsx"),
        help="另存为目标路径；未提供则覆盖原工作簿。",
    )
    return parser.parse_args()


def coerce_numeric(value: str):
    if value is None:
        return None
    text = value.strip()
    if not text:
        return None
    cleaned = text.replace(",", "")
    try:
        number = float(cleaned)
    except ValueError:
        return text
    return int(number) if number.is_integer() else number


def extract_triplet(rows: List[List[str]], row_index: int) -> List:
    try:
        raw = rows[row_index - 1][1:4]
    except IndexError as exc:
        raise ValueError(f"points.csv 缺少第 {row_index} 行或 B~D 列。") from exc
    if len(raw) < 3:
        raise ValueError(f"points.csv 第 {row_index} 行数据不足 3 列。")
    return [coerce_numeric(cell) for cell in raw]


def read_csv_pairs(csv_path: str) -> List[List]:
    with open(csv_path, newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.reader(handle))
    if len(rows) < 3:
        raise ValueError("points.csv 至少需要三行数据。")
    return [extract_triplet(rows, 2), extract_triplet(rows, 3)]


def find_first_empty_row(sheet: openpyxl.worksheet.worksheet.Worksheet, ref_col: int = 1) -> int:
    row_index = 4
    while True:
        cell = sheet.cell(row=row_index, column=ref_col)
        value = cell.value
        if value is None:
            return row_index
        if isinstance(value, str) and not value.strip():
            return row_index
        row_index += 1

def get_device_id_from_csv_path(csv_path: str) -> str:
    parent = os.path.basename(os.path.dirname(os.path.abspath(csv_path)))
    head = parent[:4]
    if head.isdigit():
        return head
    m = re.search(r"(\d{4})", parent)
    if m:
        return m.group(1)
    raise ValueError(f"无法从CSV上级目录名提取4位设备ID: {parent}")

def write_values(
        sheet: openpyxl.worksheet.worksheet.Worksheet,
        row_index: int,
        first_triplet: List,
        second_triplet: List,
        device_id: Optional[str] = None,
) -> None:
    if device_id is not None:
        sheet.cell(row=row_index, column=1, value=device_id)
    for offset, value in enumerate(first_triplet):
        sheet.cell(row=row_index, column=9 + offset, value=value)
    for offset, value in enumerate(second_triplet):
        sheet.cell(row=row_index, column=15 + offset, value=value)

def _zip_norm(p: str) -> str:
    return str(PurePosixPath(p))


def _xl_path_from_target(target: str) -> str:
    # target 可能是 "../media/image1.png" / "media/image1.png"
    p = PurePosixPath(target)
    while p.parts and p.parts[0] == "..":
        p = PurePosixPath(*p.parts[1:])
    return _zip_norm(PurePosixPath("xl") / p)


def _get_sheet_xml_path(zf: zipfile.ZipFile, sheet_name: str) -> str:
    ns_main = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    ns_r = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    ns_pkg = "http://schemas.openxmlformats.org/package/2006/relationships"

    wb_root = ET.fromstring(zf.read("xl/workbook.xml"))
    sheet_elem = None
    for sh in wb_root.findall(f".//{{{ns_main}}}sheet"):
        if sh.get("name") == sheet_name:
            sheet_elem = sh
            break
    if sheet_elem is None:
        raise ValueError(f"工作簿中不存在工作表: {sheet_name}")

    rid = sheet_elem.get(f"{{{ns_r}}}id")
    if not rid:
        raise ValueError(f"工作表 {sheet_name} 缺少 r:id")

    rels_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    for rel in rels_root.findall(f".//{{{ns_pkg}}}Relationship"):
        if rel.get("Id") == rid:
            target = rel.get("Target")
            # 通常是 "worksheets/sheet1.xml"
            return _zip_norm(PurePosixPath("xl") / target)

    raise ValueError(f"未找到工作表 {sheet_name} 的关系目标 (r:id={rid})")


def extract_column_images_to_dir(
    xlsx_path: str,
    sheet_name: str,
    column_index: int,
    out_dir: str,
) -> List[dict]:
    """
    从xlsx里提取“起始锚点在指定列”的图片，按行从上到下排序，
    保存到 out_dir 下，文件名用 1.png/2.jpg...，并返回映射列表：
    [{"row": 12, "col": 5, "file": "1.png"}, ...]
    """
    os.makedirs(out_dir, exist_ok=True)

    ns_main = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    ns_r = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    ns_pkg = "http://schemas.openxmlformats.org/package/2006/relationships"
    ns_xdr = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    ns_a = "http://schemas.openxmlformats.org/drawingml/2006/main"

    mappings: List[dict] = []

    with zipfile.ZipFile(xlsx_path, "r") as zf:
        ws_xml_path = _get_sheet_xml_path(zf, sheet_name)

        ws_root = ET.fromstring(zf.read(ws_xml_path))
        drawing_elem = ws_root.find(f".//{{{ns_main}}}drawing")
        if drawing_elem is None:
            return []

        drawing_rid = drawing_elem.get(f"{{{ns_r}}}id")
        if not drawing_rid:
            return []

        # 找到 worksheet 的 rels
        ws_file = PurePosixPath(ws_xml_path).name  # sheet1.xml
        ws_rels_path = _zip_norm(PurePosixPath("xl/worksheets/_rels") / f"{ws_file}.rels")
        ws_rels_root = ET.fromstring(zf.read(ws_rels_path))

        drawing_target = None
        for rel in ws_rels_root.findall(f".//{{{ns_pkg}}}Relationship"):
            if rel.get("Id") == drawing_rid:
                drawing_target = rel.get("Target")
                break
        if not drawing_target:
            return []

        drawing_xml_path = _xl_path_from_target(drawing_target)  # xl/drawings/drawing1.xml
        drawing_root = ET.fromstring(zf.read(drawing_xml_path))

        # drawing 的 rels：把 rId -> media 路径
        drawing_file = PurePosixPath(drawing_xml_path).name  # drawing1.xml
        drawing_rels_path = _zip_norm(PurePosixPath("xl/drawings/_rels") / f"{drawing_file}.rels")
        drawing_rels_root = ET.fromstring(zf.read(drawing_rels_path))

        embed_to_media = {}
        for rel in drawing_rels_root.findall(f".//{{{ns_pkg}}}Relationship"):
            rid = rel.get("Id")
            target = rel.get("Target")
            if rid and target and "media/" in target:
                embed_to_media[rid] = _xl_path_from_target(target)

        # 遍历 anchor
        wanted_col0 = column_index - 1  # xml里是0基
        found = []
        for tag in (f"{{{ns_xdr}}}oneCellAnchor", f"{{{ns_xdr}}}twoCellAnchor"):
            for anchor in drawing_root.findall(tag):
                from_marker = anchor.find(f"{{{ns_xdr}}}from")
                if from_marker is None:
                    continue
                col_el = from_marker.find(f"{{{ns_xdr}}}col")
                row_el = from_marker.find(f"{{{ns_xdr}}}row")
                if col_el is None or row_el is None:
                    continue
                try:
                    col0 = int(col_el.text)
                    row0 = int(row_el.text)
                except (TypeError, ValueError):
                    continue
                if col0 != wanted_col0:
                    continue

                blip = anchor.find(f".//{{{ns_a}}}blip")
                if blip is None:
                    continue
                embed = blip.get(f"{{{ns_r}}}embed")
                if not embed:
                    continue
                media_path = embed_to_media.get(embed)
                if not media_path:
                    continue

                found.append((row0 + 1, media_path))

        found.sort(key=lambda x: x[0])  # 按行从上到下

        idx = 1
        for row1, media_path in found:
            ext = os.path.splitext(media_path)[1] or ".png"
            file_name = f"{idx}{ext}"
            idx += 1
            out_path = os.path.join(out_dir, file_name)
            with open(out_path, "wb") as f:
                f.write(zf.read(media_path))
            mappings.append({"row": row1, "col": column_index, "file": file_name})

    return mappings


def _image_start_col_row(img) -> Optional[tuple]:
    # 尽量从 openpyxl image anchor 里读出起点 (col,row)，col/row 都是 1 基
    anchor = getattr(img, "anchor", None)
    if anchor is None:
        return None

    if isinstance(anchor, str):
        # 'E12'
        from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
        col_letter, row = coordinate_from_string(anchor)
        return column_index_from_string(col_letter), int(row)

    # oneCell/twoCell anchor
    from_marker = getattr(anchor, "_from", None)
    if from_marker is not None and hasattr(from_marker, "col") and hasattr(from_marker, "row"):
        return int(from_marker.col) + 1, int(from_marker.row) + 1

    return None


def remove_images_starting_in_column(ws, column_index: int) -> None:
    imgs = list(getattr(ws, "_images", []))
    kept = []
    for img in imgs:
        pos = _image_start_col_row(img)
        if not pos:
            kept.append(img)
            continue
        col, _row = pos
        if col != column_index:
            kept.append(img)
    ws._images = kept


def reinsert_column_images_from_dir(ws, mappings: List[dict], images_dir: str) -> None:
    # 依映射写回同一列同一行
    for item in mappings:
        row = int(item["row"])
        col = int(item["col"])
        img_path = os.path.join(images_dir, item["file"])

        img = XLImage(img_path)
        img.anchor = f"{get_column_letter(col)}{row}"
        ws.add_image(img)

def process_excel_with_shapes(input_path=None, sheet_name=None, output_path=None, triplets=None, device_id=None):
    """处理包含形状的Excel文件，确保形状资源不丢失"""
    with TemporaryDirectory() as temp_dir:
        # 解压原始Excel
        original_extract = os.path.join(temp_dir, "original")
        with zipfile.ZipFile(input_path, "r") as zip_ref:
            zip_ref.extractall(original_extract)

        # openpyxl编辑内容
        wb = load_workbook(input_path, keep_vba=True)

        if sheet_name not in wb.sheetnames:
            raise ValueError(f"工作簿中不存在工作表: {sheet_name}")
        ws = wb[sheet_name]
        # ===== 用户编辑区域 =====
        # sheet = wb.active
        # sheet["A1"] = "修改后仍保留所有形状资源"
        target_row = find_first_empty_row(ws)
        if device_id is not None:
            ws.cell(row=target_row, column=1, value=device_id)
        if triplets is not None:
            write_values(ws, target_row, triplets[0], triplets[1])

        # ===== 编辑结束 =====

        # 保存临时文件并解压
        temp_excel = os.path.join(temp_dir, "temp.xlsx")
        wb.save(temp_excel)
        modified_extract = os.path.join(temp_dir, "modified")
        with zipfile.ZipFile(temp_excel, "r") as zip_ref:
            zip_ref.extractall(modified_extract)

        # 关键资源恢复
        restore_shape_resources(original_extract, modified_extract)

        # 重新打包
        with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as zip_ref:
            for root, _, files in os.walk(modified_extract):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, modified_extract)
                    zip_ref.write(file_path, arcname)

    print(f"文件处理完成: {output_path}")
    print("所有形状资源已完美保留!")

    return target_row


def restore_shape_resources(orig_dir, mod_dir):
    """恢复所有形状相关资源"""
    # 1. 复制核心资源目录
    shape_dirs = [
        "xl/drawings", "xl/drawings/_rels",
        "xl/media", "xl/charts",
        "xl/embeddings", "xl/activeX"
    ]
    for dir_path in shape_dirs:
        src = os.path.join(orig_dir, dir_path)
        dest = os.path.join(mod_dir, dir_path)
        if os.path.exists(src):
            shutil.rmtree(dest, ignore_errors=True)
            shutil.copytree(src, dest)

    # 2. 修复各级关系
    fix_workbook_relationships(orig_dir, mod_dir)
    fix_worksheet_relationships(orig_dir, mod_dir)

    # 3. 恢复工作表drawing标签
    restore_drawing_tags(orig_dir, mod_dir)

    # 4. 修复内容类型声明
    fix_content_types(orig_dir, mod_dir)


def restore_drawing_tags(orig_dir, mod_dir):
    """恢复工作表级别的<drawing>标签"""
    orig_ws_dir = os.path.join(orig_dir, "xl/worksheets")
    mod_ws_dir = os.path.join(mod_dir, "xl/worksheets")

    for sheet_file in os.listdir(orig_ws_dir):
        if not sheet_file.endswith(".xml"):
            continue

        orig_path = os.path.join(orig_ws_dir, sheet_file)
        mod_path = os.path.join(mod_ws_dir, sheet_file)

        if not os.path.exists(mod_path):
            continue

        # 解析XML
        orig_tree = etree.parse(orig_path)
        mod_tree = etree.parse(mod_path)
        orig_root = orig_tree.getroot()
        mod_root = mod_tree.getroot()

        # 查找原始drawing标签
        drawing = None
        ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        for elem in orig_root.findall("main:drawing", ns):
            drawing = elem
            break

        # 如果修改后的文件缺失drawing标签
        if drawing and not mod_root.find("main:drawing", ns):
            # 插入到合理位置（通常在sheetData之后）
            sheet_data = mod_root.find("main:sheetData", ns)
            if sheet_data is not None:
                sheet_data.addnext(drawing)
            else:
                mod_root.append(drawing)

            mod_tree.write(mod_path, encoding="UTF-8", xml_declaration=True)


def fix_workbook_relationships(orig_dir, mod_dir):
    """修复工作簿级关系"""
    orig_rel = os.path.join(orig_dir, "xl/_rels/workbook.xml.rels")
    mod_rel = os.path.join(mod_dir, "xl/_rels/workbook.xml.rels")

    if not os.path.exists(orig_rel) or not os.path.exists(mod_rel):
        return

    # 解析关系文件
    ns = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
    orig_tree = ET.parse(orig_rel)
    mod_tree = ET.parse(mod_rel)
    orig_root = orig_tree.getroot()
    mod_root = mod_tree.getroot()

    # 收集现有ID
    existing_ids = {rel.get("Id") for rel in mod_root.findall("r:Relationship", ns)}

    # 添加缺失的形状关系
    for rel in orig_root.findall("r:Relationship", ns):
        rel_type = rel.get("Type", "")
        if "drawing" in rel_type or "chart" in rel_type or "image" in rel_type:
            if rel.get("Id") not in existing_ids:
                mod_root.append(rel)

    mod_tree.write(mod_rel, encoding="UTF-8", xml_declaration=True)


def fix_worksheet_relationships(orig_dir, mod_dir):
    """修复工作表级关系"""
    orig_rel_dir = os.path.join(orig_dir, "xl/worksheets/_rels")
    mod_rel_dir = os.path.join(mod_dir, "xl/worksheets/_rels")

    if not os.path.exists(orig_rel_dir):
        return

    os.makedirs(mod_rel_dir, exist_ok=True)

    for rel_file in os.listdir(orig_rel_dir):
        if not rel_file.endswith(".rels"):
            continue

        orig_path = os.path.join(orig_rel_dir, rel_file)
        mod_path = os.path.join(mod_rel_dir, rel_file)

        # 不存在则直接复制
        if not os.path.exists(mod_path):
            shutil.copy2(orig_path, mod_path)
            continue

        # 合并关系
        ns = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
        orig_tree = ET.parse(orig_path)
        mod_tree = ET.parse(mod_path)
        orig_root = orig_tree.getroot()
        mod_root = mod_tree.getroot()

        existing_ids = {rel.get("Id") for rel in mod_root.findall("r:Relationship", ns)}

        for rel in orig_root.findall("r:Relationship", ns):
            if "drawing" in rel.get("Type", "") and rel.get("Id") not in existing_ids:
                mod_root.append(rel)

        mod_tree.write(mod_path, encoding="UTF-8", xml_declaration=True)


def fix_content_types(orig_dir, mod_dir):
    """修复内容类型声明"""
    orig_ct = os.path.join(orig_dir, "[Content_Types].xml")
    mod_ct = os.path.join(mod_dir, "[Content_Types].xml")

    if not os.path.exists(orig_ct) or not os.path.exists(mod_ct):
        return

    # 解析XML
    ns = {"ct": "http://schemas.openxmlformats.org/package/2006/content-types"}
    orig_tree = ET.parse(orig_ct)
    mod_tree = ET.parse(mod_ct)
    orig_root = orig_tree.getroot()
    mod_root = mod_tree.getroot()

    # 需要添加的内容类型
    content_types = [
        "vnd.openxmlformats-officedocument.drawing+xml",
        "vnd.openxmlformats-officedocument.drawingml.chart+xml",
        "vnd.openxmlformats-officedocument.vmlDrawing",
        "image/png", "image/jpeg", "image/gif"
    ]

    # 添加缺失的类型声明
    for override in orig_root.findall("ct:Override", ns):
        if any(ct in override.get("ContentType", "") for ct in content_types):
            part_name = override.get("PartName")
            # 检查是否已存在
            if not mod_root.find(f"ct:Override[@PartName='{part_name}']", ns):
                mod_root.append(override)

    mod_tree.write(mod_ct, encoding="UTF-8", xml_declaration=True)

def _coerce_com_value(v):
    # COM 写入时 None 会清空单元格；数字/字符串都可
    return v

def write_with_com_engine(
    engine: str,
    workbook_path: str,
    sheet_name: str,
    save_path: str,
    device_id: str,
    triplets,
):
    """
    engine: "auto" / "excel" / "wps"
    triplets: [ [B2,C2,D2], [B3,C3,D3] ]
    """
    import os
    import pythoncom
    import win32com.client

    pythoncom.CoInitialize()
    app = None
    wb = None
    try:
        progids = []
        eng = (engine or "auto").lower()
        if eng == "excel":
            progids = ["Excel.Application"]
        elif eng == "wps":
            progids = ["ket.Application"]  # WPS 表格常用
        else:
            progids = ["Excel.Application", "ket.Application"]

        last_err = None
        for pid in progids:
            try:
                app = win32com.client.DispatchEx(pid)
                break
            except Exception as e:
                last_err = e
                app = None
        if app is None:
            raise RuntimeError(f"无法启动 Excel/WPS COM：{last_err}")

        app.Visible = False
        try:
            app.DisplayAlerts = False
        except Exception:
            pass

        wb = app.Workbooks.Open(os.path.abspath(workbook_path))

        try:
            ws = wb.Worksheets(sheet_name)
        except Exception:
            # 有些环境只能按索引取，再比对 Name
            ws = None
            for i in range(1, wb.Worksheets.Count + 1):
                tmp = wb.Worksheets(i)
                if tmp.Name == sheet_name:
                    ws = tmp
                    break
            if ws is None:
                raise ValueError(f"工作簿中不存在工作表: {sheet_name}")

        # 从第4行开始找A列空行
        target_row = 4
        while True:
            val = ws.Cells(target_row, 1).Value
            if val is None or (isinstance(val, str) and val.strip() == ""):
                break
            target_row += 1

        # 写设备ID到 A 列
        ws.Cells(target_row, 1).Value = device_id

        # 写 IJK (9~11)
        for offset, value in enumerate(triplets[0]):
            ws.Cells(target_row, 9 + offset).Value = _coerce_com_value(value)

        # 写 OPQ (15~17)
        for offset, value in enumerate(triplets[1]):
            ws.Cells(target_row, 15 + offset).Value = _coerce_com_value(value)

        # 保存：同路径用 Save，不同路径用 SaveAs
        src = os.path.abspath(workbook_path)
        dst = os.path.abspath(save_path) if save_path else src
        if os.path.normcase(src) == os.path.normcase(dst):
            wb.Save()
        else:
            ext = os.path.splitext(dst)[1].lower()
            # Excel 常用 FileFormat：xlsx=51, xlsm=52
            file_format = 52 if ext == ".xlsm" else 51
            try:
                wb.SaveAs(dst, FileFormat=file_format)
            except Exception:
                # WPS 有时不吃 FileFormat
                wb.SaveAs(dst)

        return target_row
    finally:
        try:
            if wb is not None:
                wb.Close(SaveChanges=False)
        except Exception:
            pass
        try:
            if app is not None:
                app.Quit()
        except Exception:
            pass
        pythoncom.CoUninitialize()

def main() -> None:
    args = parse_args()
    csv_path = os.path.abspath(args.csv)
    workbook_path = os.path.abspath(args.workbook)
    save_path = os.path.abspath(args.save_as) if args.save_as else workbook_path

    if not os.path.exists(csv_path):
        raise FileNotFoundError(f"未找到 csv 文件: {csv_path}")
    if not os.path.exists(workbook_path):
        raise FileNotFoundError(f"未找到 Excel 文件: {workbook_path}")

    triplets = read_csv_pairs(csv_path)
    device_id = get_device_id_from_csv_path(csv_path)


    if args.preserve_image_column:
        with TemporaryDirectory() as temp_dir:
            images_dir = os.path.join(os.getcwd(), "col_images")
            mappings = extract_column_images_to_dir(
                workbook_path, args.sheet, args.preserve_image_column, images_dir
            )

            wb = load_workbook(workbook_path, keep_vba=True)
            if args.sheet not in wb.sheetnames:
                raise ValueError(f"工作簿中不存在工作表: {args.sheet}")
            ws = wb[args.sheet]

            # 先把该列旧图移除（避免重复），再写入数据
            remove_images_starting_in_column(ws, args.preserve_image_column)

            target_row = find_first_empty_row(ws)
            write_values(ws, target_row, triplets[0], triplets[1], device_id=device_id)

            # 再按原行号把图插回去
            reinsert_column_images_from_dir(ws, mappings, images_dir)

            wb.save(save_path)
    else:
        target_row = process_excel_with_shapes(workbook_path, args.sheet, save_path, triplets, device_id=device_id)

    print(f"已在第 {target_row} 行写入数据并保存到: {save_path}")
    # write_values(sheet, target_row, triplets[0], triplets[1])
    try:
        target_row = write_with_com_engine(
            engine="auto",  # 或 "excel" / "wps"
            workbook_path=workbook_path,
            sheet_name=args.sheet,
            save_path=save_path,
            device_id=device_id,
            triplets=triplets,
        )
        print(f"已在第 {target_row} 行写入并保存到: {save_path}")

    except PermissionError as e:
        raise PermissionError(f"保存文件失败，请检查文件是否已被其他程序占用。") from e


if __name__ == "__main__":
    main()